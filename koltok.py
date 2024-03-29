import logging

from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

logging.basicConfig(filename='koltok.log',filemode='w', encoding='utf-8', level=logging.INFO)

link = "https://hu.wikipedia.org/wiki/Magyar_k%C3%B6lt%C5%91k,_%C3%ADr%C3%B3k_list%C3%A1ja"

lista = []
kivetelElemek = []
wb = load_workbook("koltok.xlsx")
ws = wb["költök"]
wsb = wb["kivetellista"]
excelFejlec= ["Köktő neve", "Születési helye", "Születés időpontja", "Halála ideje"]



# ---------------------------------Excel műveletel
# Excel tábla kiüritése
def excelUrites():
    logging.info("Excel tábla ürítés elindúl")
    i = 1
    while (ws['A' + str(i)].value):
        ws['A' + str(i)] = " "
        i = i + 1

# Excel tábla mentése
def excelMentes():
    wb.save("koltok.xlsx")

# Excel fejlécek beírása a táblázatba
def excelFejlecBeallitas():
    logging.info("Excel fejléc elkészítés")
    logging.info(excelFejlec)
    ws["A" + "1"] = excelFejlec[0]
    ws["B" + "1"] = excelFejlec[1]
    ws["C" + "1"] = excelFejlec[2]
    ws["D" + "1"] = excelFejlec[3]

def excelListaMentes(lista):
    print("Mentés lista elemek*******************")
    for i in lista:
        print(i)
    print(len(lista))
    for i in range(0, len(lista)-1):

        ws["A" + str(i+2)] = lista[i]

# Kivételek összeállítása
def kivetellista():
    visszLista = []
    i = 1
    while (wsb['A' + str(i)].value):
        visszLista.append(wsb["A" + str(i)].value)
        i = i + 1

    return visszLista

def szortirozo(kivetelElemek, lista):

    for i in kivetelElemek:
        if i in lista:

            lista.remove(i)


    return lista

def kovertLista(lista):
    visszaLista = []
    # for i in range(0,500):
    for i in range(0,len(lista)):
        visszaLista.append(lista[i].text)

    return visszaLista
# -------------------------------------Web műveletek
def openBrowser():
    driver = webdriver.Chrome()
    driver.get(link)
    driver.minimize_window()
    return driver

# Lista létrehozása a wiki oldalról
def koltoLista(driver):
    return driver.find_elements(By.PARTIAL_LINK_TEXT, " ")

def weblapBezarasa():
    driver.close()




logging.info("Start pogram")
excelUrites()
excelFejlecBeallitas()
driver = openBrowser()
lista = koltoLista(driver)
lista = kovertLista(lista)
kivetelElemek = kivetellista()
listaM = szortirozo(kivetelElemek, lista)
excelListaMentes(listaM)

excelMentes()