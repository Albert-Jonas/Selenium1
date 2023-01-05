import time
import logging
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook

logging.basicConfig(filename='calc_Debug.log', filemode='w', encoding='utf-8', level=logging.DEBUG)
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--lang= hu")


# A browser indítása
def startBrowser():
    driver = webdriver.Chrome()
    driver.get("https://www.desmos.com/scientific")
    return driver

# Gombok műveleti jel nyomogatása a felületen
def test(a,b,muvelet):
    for i in str(a):
        driver.find_element(By.CSS_SELECTOR, "span[aria-label='{}']".format(i)).click()
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, "span[aria-label='{}']".format(muvelet)).click()
    time.sleep(2)
    for i in str(b):
        driver.find_element(By.CSS_SELECTOR, "span[aria-label='{}']".format(i)).click()

# Eredmény beolvasása a felületről
def eredmeny():
    vissza = driver.find_elements(By.XPATH, "/html/body/div[2]/div/div[2]/div/div/div/div[1]/div[2]/div[1]/div[5]/div[2]/div[1]")
    strErtek = ""
    for Item in vissza:
        strErtek += Item.text
    return strErtek[1:]

def biralat(szam3, eredmenystr, i):
    try:
        assert str(szam3)  == eredmenystr, "Hibás"
    except AssertionError as e:
        ws['F' + str(i)] = "Fail"
        ws['G' + str(i)] = str(e)
        logging.error("Logfile entry")
    else:
        ws['F' + str(i)] = "Pass"

# Összeadás művelet
def osszeadas_pozitiv_szamokkal(szam1, szam2, szam3, i):
    time.sleep(2)
    test(szam1, szam2, "Összeadás")
    time.sleep(2)
    eredmenystr = eredmeny()
    biralat(szam3, eredmenystr,i)


# Kivonás művelet
def kivonas_pozitiv_szamokkal(szam1, szam2, szam3, i):
    time.sleep(2)
    test(szam1, szam2, "Kivonás")
    time.sleep(2)
    eredmenystr = eredmeny()
    biralat(szam3, eredmenystr, i)


# Szorzás művelet
def szorzas_pozitiv_szamokkal(szam1, szam2, szam3, i):
    time.sleep(2)
    test(szam1, szam2, "Szorzás")
    time.sleep(2)
    eredmenystr = eredmeny()
    biralat(szam3,eredmenystr, i)

# Összeadás művelet
def osztas_pozitiv_szamokkal(szam1, szam2, szam3, i):
    time.sleep(2)
    test(szam1, szam2, "Osztás")
    time.sleep(2)
    eredmenystr = eredmeny()
    biralat(szam3, eredmenystr, i)

# Kinulázza a számológép beviteli mezőt
def clear():
    driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/div/div/div/div[3]/div[1]/div/div[7]").click()


#  Excel site, cellákat alaphelyzetbe (üresre) állítja
def excelReset(ws,wb):
    i=3
    while (ws['B' + str(i)].value):
        ws['F' + str(i)] = " "
        ws['G' + str(i)] = " "
        i = i + 1

# Start
logging.error("Start")
driver = startBrowser()
# tabla , site
wb = load_workbook('excel.xlsx')
# ws = wb["Összeadás"]
muvelet = 1
# müveleti jelzö (1-4)

# Művelet választó - Site hozzárendelés
while(muvelet <=4):
    # print(muvelet)
    match muvelet:
        case 1:
            ws = wb["Összeadás"]
            excelReset(ws,wb)
        case 2:
            ws = wb["Kivonás"]
            excelReset(ws, wb)
        case 3:
            ws = wb["Szorzás"]
            excelReset(ws, wb)
        case 4:
            ws = wb["Osztás"]
            excelReset(ws, wb)

    i = 3
    # Megfelelő számolás indítésa
    while(ws['B'+str(i)].value):
        match muvelet:
            case 1:
                osszeadas_pozitiv_szamokkal(ws['B'+str(i)].value, ws['C'+str(i)].value, ws['D'+str(i)].value, i)
            case 2:
                kivonas_pozitiv_szamokkal(ws['B'+str(i)].value, ws['C'+str(i)].value, ws['D'+str(i)].value, i)
            case 3:
                szorzas_pozitiv_szamokkal(ws['B'+str(i)].value, ws['C'+str(i)].value, ws['D'+str(i)].value, i)
            case 4:
                osztas_pozitiv_szamokkal(ws['B'+str(i)].value, ws['C'+str(i)].value, ws['D'+str(i)].value, i)
        # számláló növelése
        i = i + 1
        # Beviteli mező tőrlő hívása
        clear()
    # Müvelet kiválasztó érték növelése
    muvelet = muvelet+1
# Excel tábla mentése
wb.save('excel.xlsx')
# Oldal bezárása
driver.close()
# driver.close()
